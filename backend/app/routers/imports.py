"""Importação e exportação de templates Excel para produtos e insumos."""
import io
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.ingredient import Ingredient
from app.models.product import Product
from app.models.user import User
from app.core.dependencies import require_admin

router = APIRouter(prefix="/import", tags=["Importação Excel"])

_ORANGE = "F97316"
_HEADER_FG = "FFFFFF"
_SAMPLE_BG = "FFF7ED"
_SAMPLE_FG = "999999"
_thin = Side(style="thin", color="E5E7EB")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _header_cell(ws, row: int, col: int, value: str, width: int):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color=_HEADER_FG, size=11)
    c.fill = PatternFill("solid", fgColor=_ORANGE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _border
    ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 22


def _sample_cell(ws, row: int, col: int, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(color=_SAMPLE_FG, italic=True)
    c.fill = PatternFill("solid", fgColor=_SAMPLE_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border


def _excel_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── INSUMOS ─────────────────────────────────────────────────────────────────

_ING_COLS = [
    ("nome *", 30),
    ("unidade_de_medida *", 20),
    ("custo_unitario *", 18),
    ("codigo", 15),
    ("categoria", 20),
    ("fornecedor", 25),
    ("observacoes", 35),
]
_ING_SAMPLE = ["Farinha de Trigo", "kg", "3.50", "FT001", "Secos", "Distribuidora X", "Armazenar em local seco"]
_UNITS = "kg,g,L,ml,un,cx,pct,ds,lt,sc"


@router.get("/template/ingredients", summary="Baixar modelo Excel de insumos")
async def template_ingredients(current_user: User = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Insumos"

    # Instrução
    ws.merge_cells(f"A1:{get_column_letter(len(_ING_COLS))}1")
    ws["A1"] = (
        "MODELO DE IMPORTAÇÃO — INSUMOS | "
        "Preencha a partir da linha 3. Não altere os cabeçalhos. "
        "A linha 2 é apenas exemplo — apague ou substitua."
    )
    ws["A1"].font = Font(italic=True, color="555555", size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for col_idx, (header, width) in enumerate(_ING_COLS, 1):
        _header_cell(ws, 2, col_idx, header, width)

    for col_idx, val in enumerate(_ING_SAMPLE, 1):
        _sample_cell(ws, 3, col_idx, val)

    # Dropdown para unidade_de_medida
    dv = DataValidation(type="list", formula1=f'"{_UNITS}"', showDropDown=False, showErrorMessage=True)
    dv.error = f"Use uma das unidades: {_UNITS}"
    dv.errorTitle = "Unidade inválida"
    ws.add_data_validation(dv)
    dv.sqref = "B3:B5000"

    ws.freeze_panes = "A3"
    return _excel_response(wb, "modelo_insumos.xlsx")


@router.post("/ingredients", summary="Importar insumos via Excel")
async def import_ingredients(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Linha vazia
        if not any(cell for cell in row if cell is not None and str(cell).strip()):
            skipped += 1
            continue

        name = str(row[0]).strip() if row[0] is not None else ""
        unit = str(row[1]).strip() if row[1] is not None else ""
        cost_raw = row[2]
        code = str(row[3]).strip() if row[3] is not None else None
        category = str(row[4]).strip() if row[4] is not None else None
        supplier = str(row[5]).strip() if row[5] is not None else None
        notes = str(row[6]).strip() if row[6] is not None else None

        if not name:
            errors.append({"linha": row_idx, "erro": "Nome é obrigatório"})
            continue
        if not unit:
            errors.append({"linha": row_idx, "erro": f"'{name}': unidade é obrigatória"})
            continue

        try:
            unit_cost = Decimal(str(cost_raw).replace(",", ".")).quantize(Decimal("0.0001")) if cost_raw is not None else Decimal("0")
            if unit_cost < 0:
                raise ValueError()
        except (InvalidOperation, ValueError):
            errors.append({"linha": row_idx, "erro": f"'{name}': custo inválido → '{cost_raw}'"})
            continue

        db.add(Ingredient(
            tenant_id=current_user.tenant_id,
            name=name,
            unit_of_measure=unit,
            unit_cost=unit_cost,
            code=code or None,
            category=category or None,
            supplier=supplier or None,
            notes=notes or None,
        ))
        created += 1

    await db.flush()
    return {"criados": created, "linhas_vazias_ignoradas": skipped, "erros": errors}


# ─── PRODUTOS ────────────────────────────────────────────────────────────────

_PROD_COLS = [
    ("nome *", 35),
    ("codigo", 15),
    ("preco_de_venda", 18),
    ("curva_abc", 12),
    ("descricao", 40),
    ("modo_de_preparo", 50),
]
_PROD_SAMPLE = ["X-Burguer Artesanal", "XB001", "29.90", "A", "Hambúrguer artesanal 180g", "Grelhar o blend por 4min de cada lado..."]


@router.get("/template/products", summary="Baixar modelo Excel de produtos")
async def template_products(current_user: User = Depends(require_admin)):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produtos"

    ws.merge_cells(f"A1:{get_column_letter(len(_PROD_COLS))}1")
    ws["A1"] = (
        "MODELO DE IMPORTAÇÃO — PRODUTOS | "
        "Preencha a partir da linha 3. Não altere os cabeçalhos. "
        "A linha 2 é apenas exemplo — apague ou substitua."
    )
    ws["A1"].font = Font(italic=True, color="555555", size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for col_idx, (header, width) in enumerate(_PROD_COLS, 1):
        _header_cell(ws, 2, col_idx, header, width)

    for col_idx, val in enumerate(_PROD_SAMPLE, 1):
        _sample_cell(ws, 3, col_idx, val)

    # Dropdown para curva_abc
    dv = DataValidation(type="list", formula1='"A,B,C"', showDropDown=False, showErrorMessage=True)
    dv.error = "Use A, B ou C"
    dv.errorTitle = "Curva inválida"
    ws.add_data_validation(dv)
    dv.sqref = "D3:D5000"

    ws.freeze_panes = "A3"
    return _excel_response(wb, "modelo_produtos.xlsx")


@router.post("/products", summary="Importar produtos via Excel")
async def import_products(
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    created = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(cell for cell in row if cell is not None and str(cell).strip()):
            skipped += 1
            continue

        name = str(row[0]).strip() if row[0] is not None else ""
        code = str(row[1]).strip() if row[1] is not None else None
        price_raw = row[2]
        abc_raw = str(row[3]).strip().upper() if row[3] is not None else None
        description = str(row[4]).strip() if row[4] is not None else None
        preparation = str(row[5]).strip() if row[5] is not None else None

        if not name:
            errors.append({"linha": row_idx, "erro": "Nome é obrigatório"})
            continue

        selling_price = None
        if price_raw is not None and str(price_raw).strip():
            try:
                selling_price = Decimal(str(price_raw).replace(",", ".")).quantize(Decimal("0.01"))
                if selling_price < 0:
                    raise ValueError()
            except (InvalidOperation, ValueError):
                errors.append({"linha": row_idx, "erro": f"'{name}': preço inválido → '{price_raw}'"})
                continue

        abc_curve = None
        if abc_raw in ("A", "B", "C"):
            abc_curve = abc_raw
        elif abc_raw:
            errors.append({"linha": row_idx, "erro": f"'{name}': curva_abc inválida → '{abc_raw}' (use A, B ou C)"})
            continue

        db.add(Product(
            tenant_id=current_user.tenant_id,
            name=name,
            code=code or None,
            selling_price=selling_price,
            abc_curve=abc_curve,
            description=description or None,
            preparation_method=preparation or None,
        ))
        created += 1

    await db.flush()
    return {"criados": created, "linhas_vazias_ignoradas": skipped, "erros": errors}
